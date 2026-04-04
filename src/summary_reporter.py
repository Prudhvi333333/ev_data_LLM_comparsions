from __future__ import annotations

from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from src.utils.logger import get_logger


logger = get_logger("summary_reporter")

SUMMARY_METRICS = [
    "faithfulness",
    "answer_relevancy",
    "context_precision",
    "context_recall",
    "answer_correctness",
    "final_score",
]


def _read_summary_metrics(report_path: str) -> dict[str, float]:
    workbook = load_workbook(report_path, data_only=True)
    sheet = workbook["Summary"]
    values: dict[str, float] = {}
    for row_idx in range(2, sheet.max_row + 1):
        metric_name = str(sheet.cell(row=row_idx, column=1).value or "").strip()
        if metric_name:
            values[metric_name] = float(sheet.cell(row=row_idx, column=2).value or 0.0)
    return values


def _apply_extreme_fills(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red = PatternFill(fill_type="solid", fgColor="F4CCCC")
    for col_idx in range(start_col, end_col + 1):
        values = [
            float(sheet.cell(row=row_idx, column=col_idx).value or 0.0)
            for row_idx in range(start_row, end_row + 1)
        ]
        if not values:
            continue
        max_value = max(values)
        min_value = min(values)
        for row_idx in range(start_row, end_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.number_format = "0.0000"
            numeric_value = float(cell.value or 0.0)
            if numeric_value == max_value:
                cell.fill = green
            if numeric_value == min_value:
                cell.fill = red


def _style_header(sheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def build_cross_pipeline_summary(report_paths: dict[str, str], config: SimpleNamespace) -> str:
    rows = []
    for pipeline_id, report_path in report_paths.items():
        if not Path(report_path).exists():
            continue
        summary_values = _read_summary_metrics(report_path)
        rows.append(
            {
                "pipeline": pipeline_id,
                **{metric: summary_values.get(metric, 0.0) for metric in SUMMARY_METRICS},
            }
        )

    workbook = Workbook()
    all_sheet = workbook.active
    all_sheet.title = "All Pipelines"
    all_sheet.append(["Pipeline", *SUMMARY_METRICS])
    for row in rows:
        all_sheet.append([row["pipeline"], *[row[metric] for metric in SUMMARY_METRICS]])
    _style_header(all_sheet)
    if rows:
        _apply_extreme_fills(all_sheet, 2, len(rows) + 1, 2, len(SUMMARY_METRICS) + 1)

    rag_vs_norag = workbook.create_sheet("RAG vs No-RAG")
    rag_vs_norag.append(["Model", "RAG Final Score", "No-RAG Final Score", "Delta"])
    _style_header(rag_vs_norag)
    for model_key in ("qwen", "gemma", "gemini"):
        rag_score = next((row["final_score"] for row in rows if row["pipeline"] == f"{model_key}_rag"), 0.0)
        norag_score = next((row["final_score"] for row in rows if row["pipeline"] == f"{model_key}_norag"), 0.0)
        rag_vs_norag.append([model_key, rag_score, norag_score, rag_score - norag_score])
    _apply_extreme_fills(rag_vs_norag, 2, rag_vs_norag.max_row, 2, 4)

    ranking_sheet = workbook.create_sheet("Rankings")
    ranking_sheet.append(["Rank", "Pipeline", "Final Score"])
    _style_header(ranking_sheet)
    ranked_rows = sorted(rows, key=lambda row: row["final_score"], reverse=True)
    for rank, row in enumerate(ranked_rows, start=1):
        ranking_sheet.append([rank, row["pipeline"], row["final_score"]])
    if ranked_rows:
        _apply_extreme_fills(ranking_sheet, 2, ranking_sheet.max_row, 3, 3)

    for sheet in (all_sheet, rag_vs_norag, ranking_sheet):
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 40)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(config.paths.output) / f"FINAL_COMPARISON_{timestamp}.xlsx"
    workbook.save(output_path)
    logger.info("Saved cross-pipeline summary %s", output_path)
    return str(output_path)
