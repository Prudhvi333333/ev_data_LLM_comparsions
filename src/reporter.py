from __future__ import annotations

from datetime import datetime
from pathlib import Path
from statistics import mean, pstdev
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from src.utils.config_loader import get_pipeline_id
from src.utils.logger import get_logger


logger = get_logger("reporter")

RESULT_HEADERS = [
    "Question_ID",
    "Use_Case_Category",
    "Question",
    "Golden_Answer",
    "Generated_Answer",
    "Retrieved_Context",
    "Faithfulness",
    "Answer_Relevancy",
    "Context_Precision",
    "Context_Recall",
    "Answer_Correctness",
    "Final_Score",
    "Faithfulness_Reason",
    "Relevancy_Reason",
]

METRIC_KEYS = [
    "faithfulness",
    "answer_relevancy",
    "context_precision",
    "context_recall",
    "answer_correctness",
    "final_score",
]


def _score_fill(value: float) -> PatternFill:
    if value >= 0.7:
        return PatternFill(fill_type="solid", fgColor="C6EFCE")
    if value >= 0.5:
        return PatternFill(fill_type="solid", fgColor="FFF2CC")
    return PatternFill(fill_type="solid", fgColor="F4CCCC")


def _autofit_sheet_columns(sheet, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def _apply_readable_column_widths(sheet) -> None:
    width_map = {
        "A": 12,
        "B": 34,
        "C": 50,
        "D": 60,
        "E": 80,
        "F": 90,
        "M": 50,
        "N": 50,
    }
    for column_letter, width in width_map.items():
        sheet.column_dimensions[column_letter].width = width


def _style_results_sheet(sheet, row_count: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    light_blue = PatternFill(fill_type="solid", fgColor="D9EAF7")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    white_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, row_count + 2):
        sheet.row_dimensions[row_idx].height = 90
        fill = light_blue if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, len(RESULT_HEADERS) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(7, 13):
            cell = sheet.cell(row=row_idx, column=col_idx)
            numeric_value = float(cell.value or 0.0)
            cell.number_format = "0.0000"
            cell.fill = _score_fill(numeric_value)

    sheet.freeze_panes = "A2"


def _build_summary_sheet(workbook: Workbook, results: list[dict[str, Any]]) -> None:
    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Mean", "Std Dev"])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    for metric in METRIC_KEYS:
        values = [float(result.get(metric, 0.0) or 0.0) for result in results]
        metric_mean = mean(values) if values else 0.0
        metric_std = pstdev(values) if len(values) > 1 else 0.0
        summary.append([metric, metric_mean, metric_std])

    for row_idx in range(2, summary.max_row + 1):
        for col_idx in (2, 3):
            cell = summary.cell(row=row_idx, column=col_idx)
            numeric_value = float(cell.value or 0.0)
            cell.number_format = "0.0000"
            if col_idx == 2:
                cell.fill = _score_fill(numeric_value)

    chart = BarChart()
    chart.title = "Mean Metric Scores"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Metric"
    data = Reference(summary, min_col=2, min_row=1, max_row=summary.max_row)
    categories = Reference(summary, min_col=1, min_row=2, max_row=summary.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 14
    summary.add_chart(chart, "E2")

    _autofit_sheet_columns(summary)


def build_report(results: list[dict[str, Any]], model_name: str, mode: str, config: SimpleNamespace) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(RESULT_HEADERS)

    for result in results:
        sheet.append(
            [
                result.get("question_id", ""),
                result.get("category", ""),
                result.get("question", ""),
                result.get("golden_answer", ""),
                result.get("generated_answer", ""),
                result.get("retrieved_context", ""),
                float(result.get("faithfulness", 0.0) or 0.0),
                float(result.get("answer_relevancy", 0.0) or 0.0),
                float(result.get("context_precision", 0.0) or 0.0),
                float(result.get("context_recall", 0.0) or 0.0),
                float(result.get("answer_correctness", 0.0) or 0.0),
                float(result.get("final_score", 0.0) or 0.0),
                result.get("faithfulness_reason", ""),
                result.get("answer_relevancy_reason", ""),
            ]
        )

    _style_results_sheet(sheet, len(results))
    _autofit_sheet_columns(sheet)
    _apply_readable_column_widths(sheet)
    _build_summary_sheet(workbook, results)

    report_dir = Path(config.paths.output)
    report_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = report_dir / f"{get_pipeline_id(model_name, mode)}_report_{timestamp}.xlsx"
    workbook.save(report_path)

    logger.info("Saved report %s", report_path)
    return str(report_path)
