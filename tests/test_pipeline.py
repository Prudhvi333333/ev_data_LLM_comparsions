from __future__ import annotations

import asyncio
from pathlib import Path
import sys
import unittest
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
VENDOR = ROOT / ".vendor"
if VENDOR.exists() and str(VENDOR) not in sys.path:
    sys.path.insert(0, str(VENDOR))

from openpyxl import load_workbook

from src.evaluator import RAGASEvaluator
from src.generator import OllamaGenerator
from src.indexer import get_or_build_index
from src.kb_loader import build_document_text, load_kb
from src.reporter import build_report
from src.retriever import HybridRetriever
from src.utils.async_helpers import create_timeout_guard
from src.utils.config_loader import load_config


class PipelineTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.config = load_config("config/config.yaml")
        cls.documents = load_kb(cls.config)
        cls.indexer = get_or_build_index(cls.config, cls.documents, force_reindex=True)
        cls.retriever = HybridRetriever(cls.indexer, cls.config)

    def test_kb_loader(self) -> None:
        self.assertEqual(len(self.documents), 205)
        self.assertTrue(all(doc["metadata"].get("Company") for doc in self.documents))

    def test_document_text(self) -> None:
        text = build_document_text(self.documents[0]["metadata"])
        for label in [
            "Company:",
            "Tier:",
            "Industry:",
            "Location:",
            "Address:",
            "Facility:",
            "EV Role:",
            "OEMs:",
            "Affiliation:",
            "Employment:",
            "Products:",
            "Classification:",
        ]:
            self.assertIn(label, text)

    def test_retriever_smoke(self) -> None:
        results = self.retriever.retrieve("Tier 1 battery companies")
        self.assertGreaterEqual(len(results), 3)
        self.assertTrue(all("metadata" in doc for doc in results))

    def test_intent_detection(self) -> None:
        list_intent = self.retriever.detect_query_intent("List all Tier 1 suppliers")
        self.assertTrue(list_intent["has_tier_filter"])
        self.assertTrue(list_intent["is_list_query"])

        count_intent = self.retriever.detect_query_intent(
            "How many battery cell companies are in Georgia?"
        )
        self.assertTrue(count_intent["is_count_query"])
        self.assertIn("battery cell", str(count_intent["role_value"]).lower())

    def test_generator_timeout(self) -> None:
        generator = OllamaGenerator("qwen2.5:14b", self.config)

        async def slow_call(payload, timeout_seconds):
            del payload, timeout_seconds
            await asyncio.sleep(70)
            return "slow"

        async def run_timeout_check():
            with patch.object(generator, "_invoke_ollama", slow_call):
                return await create_timeout_guard(
                    generator.generate("Question?", None),
                    timeout_sec=0.05,
                    fallback_value="TIMEOUT",
                )

        self.assertEqual(asyncio.run(run_timeout_check()), "TIMEOUT")

    def test_evaluator_json_parse(self) -> None:
        self.config.api_keys.openrouter = "test-openrouter-key"
        evaluator = RAGASEvaluator(self.config)

        async def bad_judge(prompt):
            del prompt
            return "this is not json"

        async def run_parse_check():
            with patch.object(evaluator, "_call_judge_api", bad_judge):
                return await evaluator.score_metric(
                    "faithfulness",
                    "Question",
                    "Golden answer",
                    "Generated answer",
                    "Context",
                )

        with self.assertRaises(RuntimeError):
            asyncio.run(run_parse_check())

    def test_reporter_creates_file(self) -> None:
        rows = []
        for idx in range(1, 4):
            rows.append(
                {
                    "question_id": idx,
                    "category": "Demo",
                    "question": f"Question {idx}",
                    "golden_answer": f"Golden {idx}",
                    "generated_answer": f"Generated {idx}",
                    "retrieved_context": f"Context {idx}",
                    "faithfulness": 0.8,
                    "answer_relevancy": 0.7,
                    "context_precision": 0.9,
                    "context_recall": 0.6,
                    "answer_correctness": 0.5,
                    "final_score": 0.7,
                    "faithfulness_reason": "ok",
                    "answer_relevancy_reason": "ok",
                }
            )

        report_path = build_report(rows, "qwen", "rag", self.config)
        self.assertTrue(Path(report_path).exists())

        workbook = load_workbook(report_path)
        self.assertEqual(set(workbook.sheetnames), {"Results", "Summary"})


if __name__ == "__main__":
    unittest.main()
